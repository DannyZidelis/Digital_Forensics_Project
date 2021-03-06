import argparse
from PIL import Image
from PIL.ExifTags import TAGS
import os
import hashlib
import PIL
import openpyxl
import xlsxwriter
import xlrd

def Main():
    filepath = os.path.dirname(os.path.abspath(__file__))
    if os.path.isfile('Evidence.xlsx'):
        os.remove("Evidence.xlsx")
    workbook = xlsxwriter.Workbook('Evidence.xlsx')
    worksheet = workbook.add_worksheet()

    worksheet.write(0,0, "Photo Name")
    worksheet.write(0,1, "GPSInfo")
    worksheet.write(0,2, "Make")
    worksheet.write(0,3, "Model")
    worksheet.write(0,4, "DateTime")
    worksheet.write(0,5, "ImageUniqueID")
    fullfilepath = filepath + "\\" + "Evidence.xlsx"

    row=1
    col=1
    for f in os.listdir('.'):
            #checks if the file has an image extension
            if f.lower().endswith(('.png', '.jpg', '.jpeg')):
                    metaData = {}
                    i = Image.open(f)
                    #splits the file name into name, and extension and saves to fn
                    fn, fext = os.path.splitext(f)
                    #creates name of the outfile for the specific image
                    outfile = fn +".txt"
                    #deletes text file for specific file so that it does not just append
                    #to the end of the text file
                    if os.path.exists(outfile):
                            os.remove(outfile)
                    print("Getting meta data... of ", f)
                    info = i._getexif()
                    #loop to enter in the iGPSInfo	{0: b'\x02\x02\x00\x00', 18: 'WGS-84', 5: b'\x00'}nformation
                    if info:
                            #print("found meta data")
                            for (tag, value) in info.items():
                                #print("col: ", col, "row: ", row)
                                tagname = TAGS.get(tag, tag)
                                metaData[tagname] = value
                                worksheet.write(row, col, fn)
                                col=2
                            #loops through and appends the tags to the text file
                            with open(outfile, 'a+') as f:
                                f.write(fn + "\n")
                                worksheet.write(row, 0, fn)
                                for(tagname, value) in metaData.items():
                                        f.write(str(tagname)+"\t"+str(value)+"\n")
                                        if tagname == "GPSInfo":
                                            worksheet.write(row,1, str(value))
                                            #print("gpsbro at col ",1, " row: ", row, " -- ", str(value))
                                        elif tagname == "Make":
                                            worksheet.write(row,2, value)
                                            #print("gpsbro at col ",2, " row: ", row, " -- ", value)
                                        elif tagname == "Model":
                                            worksheet.write(row,3, value)
                                            #print("gpsbro at col ",3, " row: ", row, " -- ", value)
                                        elif tagname == "DateTime":
                                            worksheet.write(row,4, value)
                                            #print("gpsbro at col ",4, " row: ", row, " -- ", value)
                                        elif tagname == "ImageUniqueID":
                                            worksheet.write(row,5, value)
                                            #print("gpsbro at col ",5, " row: ", row, " -- ", value)
                            row=row+1

    workbook.close()
    wrkbook = xlrd.open_workbook(fullfilepath)
    wrksheet = wrkbook.sheet_by_name('Sheet1')
    datefile = "dates.txt"
    if os.path.isfile(datefile):
        os.remove(datefile)
    locfile = "locations.txt"
    if os.path.isfile(locfile):
        os.remove(locfile)

    #when no value it gets IndexError
    #value = wrksheet.cell(20, 10)
    #print(value)
    #2008:11:01 21:15:11
    search_date = input("Would you like to search by date? (y or n) ")
    if search_date =='y':
        date = input("What date would you like to search for?( ****:**:** year:month:day) ")
        # i=0
        # rowcount = 1
        # while i == 0:
        #     value = wrksheet.cell(rowcount, 4)
        #     #print(value)
        #     if date not in str(value):
        #          continue
        #     else:
        #         date_name = wrksheet.cell(rowcount, 0)
        #         date_gps = wrksheet.cell(rowcount, 1)
        #         f.write(str(date_name)+"\t"+str(date_gps)+"\t"+str(value)+"\n")
        #         print(str(date_name)+"\t"+str(date_gps)+"\t"+str(value)+"\n")

    search_loc = input("Would you like to search by location? (y or n) ")
    if search_loc =='y':
        loc = input("What location would you like to search for? ")
if __name__ == '__main__':
    Main()
